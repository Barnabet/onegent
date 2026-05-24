"""
`xlsx` domain — read, inspect, query, write, edit, format, convert, and
recalculate spreadsheet files (.xlsx / .xlsm / .csv / .tsv).

Design notes
------------
- Every tool returns a ``ToolResult`` envelope; no exceptions escape.
- Heavy deps (``openpyxl``, system ``soffice``) are imported / probed lazily.
- Multi-sheet workbooks are first-class: ``xlsx.read`` exposes a ``sheets`` list,
  ``xlsx.info`` enumerates every sheet, ``xlsx.write`` and ``xlsx.sql`` accept
  a dict-of-sheets, and ``xlsx.convert`` can explode a workbook into one CSV
  per sheet.
- The ``xlsx.sql`` tool loads every sheet/CSV into an in-memory SQLite database
  and runs read-only SQL against it. This is the right tool any time the agent
  needs *computation* (aggregations, joins, filtering) instead of dumping rows.
- Number / date types from openpyxl are preserved; rows are returned as lists
  of Python primitives (str/int/float/bool/None plus ISO-format strings for
  dates and datetimes).
"""

from __future__ import annotations

import csv
import datetime as _dt
import re
import shutil
import sqlite3
import subprocess
import tempfile
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from runtime.tool_registry import ToolCtx, ToolError, ToolResult


# ---------------------------------------------------------------------------
# Shared helpers
# ---------------------------------------------------------------------------


_TABULAR_SUFFIXES = (".xlsx", ".xlsm", ".csv", ".tsv")
_XLSX_SUFFIXES = (".xlsx", ".xlsm")
_DELIMITED_SUFFIXES = (".csv", ".tsv")

# A conservative identifier rule for SQLite table names. Sheet names with
# spaces or punctuation are sanitised; the original name is kept in the
# returned ``sheet_map``.
_IDENT_RE = re.compile(r"[^A-Za-z0-9_]+")


def _err(code: str, message: str, retriable: bool = False) -> ToolResult:
    return ToolResult(ok=False, error=ToolError(code=code, message=message, retriable=retriable))


def _require_input(path_str: str, *, suffixes: Tuple[str, ...] = _TABULAR_SUFFIXES) -> Tuple[Optional[Path], Optional[ToolResult]]:
    p = Path(path_str)
    if not p.is_file():
        return None, _err("file_not_found", f"No file at {path_str!r}.")
    if p.suffix.lower() not in suffixes:
        return None, _err(
            "unsupported_format",
            f"{p.suffix!r} is not supported. Allowed: {', '.join(suffixes)}.",
        )
    return p, None


def _require_openpyxl() -> Tuple[Any, Optional[ToolResult]]:
    try:
        import openpyxl  # type: ignore
        return openpyxl, None
    except ImportError:
        return None, _err("dependency_missing", "openpyxl is not installed.")


def _delimiter_for(path: Path) -> str:
    return "\t" if path.suffix.lower() == ".tsv" else ","


def _jsonable(value: Any) -> Any:
    """Coerce openpyxl cell values into JSON-safe primitives."""
    if value is None or isinstance(value, (bool, int, float, str)):
        return value
    if isinstance(value, (_dt.datetime, _dt.date, _dt.time)):
        return value.isoformat()
    if isinstance(value, _dt.timedelta):
        return value.total_seconds()
    return str(value)


def _check_output_path(path_str: str, overwrite: bool, *, suffixes: Tuple[str, ...]) -> Tuple[Optional[Path], Optional[ToolResult]]:
    p = Path(path_str)
    if p.suffix.lower() not in suffixes:
        return None, _err(
            "invalid_input",
            f"Output {p.suffix!r} not allowed. Use one of {', '.join(suffixes)}.",
        )
    if p.exists() and not overwrite:
        return None, _err(
            "output_exists",
            f"{path_str!r} already exists. Pass overwrite=true to replace it.",
        )
    p.parent.mkdir(parents=True, exist_ok=True)
    return p, None


def _sanitise_ident(name: str, taken: set) -> str:
    cleaned = _IDENT_RE.sub("_", name).strip("_") or "sheet"
    if cleaned[0].isdigit():
        cleaned = "s_" + cleaned
    base = cleaned
    i = 2
    while cleaned.lower() in taken:
        cleaned = f"{base}_{i}"
        i += 1
    taken.add(cleaned.lower())
    return cleaned


# ---------------------------------------------------------------------------
# xlsx.read — return raw rows of a single sheet (or every sheet)
# ---------------------------------------------------------------------------


def read(params, ctx: ToolCtx) -> ToolResult:
    src, err = _require_input(params.path)
    if err:
        return err

    suffix = src.suffix.lower()
    if suffix in _XLSX_SUFFIXES:
        return _read_xlsx(src, params)
    return _read_delimited(src, params)


def _read_xlsx(path: Path, params) -> ToolResult:
    openpyxl, err = _require_openpyxl()
    if err:
        return err
    try:
        wb = openpyxl.load_workbook(filename=str(path), read_only=True, data_only=True)
    except Exception as e:
        return _err("unsupported_format", f"Failed to open as xlsx: {e}")

    max_rows = getattr(params, "max_rows", None)
    if params.all_sheets:
        sheets_out = []
        for sheet_name in wb.sheetnames:
            sheets_out.append(_dump_sheet(wb[sheet_name], sheet_name, params.has_header, max_rows))
        return ToolResult(ok=True, data={
            "sheets": sheets_out,
            "sheet_names": list(wb.sheetnames),
        })

    sheet_name = params.sheet or wb.sheetnames[0]
    if sheet_name not in wb.sheetnames:
        return _err(
            "sheet_not_found",
            f"Sheet {sheet_name!r} not in workbook. Available: {wb.sheetnames}.",
        )
    payload = _dump_sheet(wb[sheet_name], sheet_name, params.has_header, max_rows)
    payload["sheet_names"] = list(wb.sheetnames)
    return ToolResult(ok=True, data=payload)


def _dump_sheet(ws, sheet_name: str, has_header: bool, max_rows: Optional[int] = None) -> dict:
    rows: List[List[Any]] = []
    for row in ws.iter_rows(values_only=True):
        rows.append([_jsonable(v) for v in row])

    headers: Optional[List[Any]] = None
    data_rows = rows
    if has_header and rows:
        headers = rows[0]
        data_rows = rows[1:]

    total = len(data_rows)
    truncated = False
    if max_rows is not None and max_rows >= 0 and total > max_rows:
        data_rows = data_rows[:max_rows]
        truncated = True

    payload = {
        "sheet": sheet_name,
        "headers": headers,
        "rows": data_rows,
        "row_count": total,
        "col_count": len(headers) if headers is not None else (len(rows[0]) if rows else 0),
    }
    if truncated:
        payload["returned_row_count"] = len(data_rows)
        payload["truncated"] = True
        payload["truncation_note"] = (
            f"Showing first {len(data_rows)} of {total} data rows. "
            "Raise `max_rows` to see more, or use xlsx.sql for aggregations."
        )
    return payload


def _read_delimited(path: Path, params) -> ToolResult:
    delim = _delimiter_for(path)
    try:
        with path.open("r", encoding="utf-8", newline="") as f:
            rows = [row for row in csv.reader(f, delimiter=delim)]
    except UnicodeDecodeError as e:
        return _err("decode_error", f"File is not UTF-8: {e}")

    headers: Optional[List[Any]] = None
    data_rows = rows
    if params.has_header and rows:
        headers = rows[0]
        data_rows = rows[1:]

    total = len(data_rows)
    truncated = False
    max_rows = getattr(params, "max_rows", None)
    if max_rows is not None and max_rows >= 0 and total > max_rows:
        data_rows = data_rows[:max_rows]
        truncated = True

    payload = {
        "sheet": path.stem,
        "headers": headers,
        "rows": data_rows,
        "row_count": total,
        "col_count": len(headers) if headers is not None else (len(rows[0]) if rows else 0),
        "sheet_names": [path.stem],
    }
    if truncated:
        payload["returned_row_count"] = len(data_rows)
        payload["truncated"] = True
        payload["truncation_note"] = (
            f"Showing first {len(data_rows)} of {total} data rows. "
            "Raise `max_rows` to see more, or use xlsx.sql for aggregations."
        )
    return ToolResult(ok=True, data=payload)


# ---------------------------------------------------------------------------
# xlsx.info — sheet inventory + shape, no row data
# ---------------------------------------------------------------------------


def info(params, ctx: ToolCtx) -> ToolResult:
    src, err = _require_input(params.path)
    if err:
        return err

    if src.suffix.lower() in _DELIMITED_SUFFIXES:
        rows, headers = _delimited_shape(src)
        return ToolResult(ok=True, data={
            "format": src.suffix.lower().lstrip("."),
            "sheets": [{
                "name": src.stem,
                "row_count": rows,
                "col_count": len(headers) if headers else 0,
                "headers_preview": headers,
            }],
            "sheet_count": 1,
        })

    openpyxl, derr = _require_openpyxl()
    if derr:
        return derr
    try:
        wb = openpyxl.load_workbook(filename=str(src), read_only=True, data_only=True)
    except Exception as e:
        return _err("unsupported_format", f"Failed to open as xlsx: {e}")

    sheets = []
    for name in wb.sheetnames:
        ws = wb[name]
        headers_preview: List[Any] = []
        row_count = 0
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                headers_preview = [_jsonable(v) for v in row]
            row_count = i + 1
        sheets.append({
            "name": name,
            "row_count": max(row_count - 1, 0),  # excludes header
            "col_count": len(headers_preview),
            "headers_preview": headers_preview,
        })

    props = getattr(wb, "properties", None)
    metadata = {
        "title": getattr(props, "title", None),
        "creator": getattr(props, "creator", None),
        "created": _jsonable(getattr(props, "created", None)),
        "modified": _jsonable(getattr(props, "modified", None)),
    } if props else {}

    return ToolResult(ok=True, data={
        "format": src.suffix.lower().lstrip("."),
        "sheets": sheets,
        "sheet_count": len(sheets),
        "metadata": metadata,
    })


def _delimited_shape(path: Path) -> Tuple[int, Optional[List[str]]]:
    delim = _delimiter_for(path)
    with path.open("r", encoding="utf-8", newline="") as f:
        reader = csv.reader(f, delimiter=delim)
        headers: Optional[List[str]] = None
        count = 0
        for i, row in enumerate(reader):
            if i == 0:
                headers = row
                continue
            count += 1
    return count, headers


# ---------------------------------------------------------------------------
# xlsx.sql — read-only SQL over one or more table files
# ---------------------------------------------------------------------------


_FORBIDDEN_SQL = re.compile(
    r"\b(INSERT|UPDATE|DELETE|DROP|ALTER|CREATE|REPLACE|ATTACH|DETACH|"
    r"PRAGMA|REINDEX|VACUUM|BEGIN|COMMIT|ROLLBACK|SAVEPOINT)\b",
    re.IGNORECASE,
)


def sql(params, ctx: ToolCtx) -> ToolResult:
    query = (params.query or "").strip().rstrip(";")
    if not query:
        return _err("invalid_input", "query is empty.")
    if _FORBIDDEN_SQL.search(query):
        return _err(
            "forbidden_statement",
            "xlsx.sql is read-only: only SELECT / WITH queries are allowed.",
        )
    if ";" in query:
        return _err("invalid_input", "Multiple statements are not allowed; submit one query.")
    leading = query.lstrip().split(None, 1)[0].upper()
    if leading not in ("SELECT", "WITH", "VALUES"):
        return _err(
            "forbidden_statement",
            f"Query must start with SELECT, WITH, or VALUES (got {leading!r}).",
        )

    # Validate inputs
    if not params.inputs:
        return _err("invalid_input", "At least one input path is required.")

    conn = sqlite3.connect(":memory:")
    sheet_map: Dict[str, dict] = {}
    taken_idents: set = set()

    try:
        for spec in params.inputs:
            if isinstance(spec, str):
                path_str, alias, sheet_filter = spec, None, None
            else:
                # Either a dict or a pydantic SqlInputSpec — both expose attributes via getattr/get.
                path_str = getattr(spec, "path", None) if not isinstance(spec, dict) else spec["path"]
                alias = getattr(spec, "alias", None) if not isinstance(spec, dict) else spec.get("alias")
                sheet_filter = getattr(spec, "sheet", None) if not isinstance(spec, dict) else spec.get("sheet")

            src, err = _require_input(path_str)
            if err:
                return err

            suffix = src.suffix.lower()
            if suffix in _DELIMITED_SUFFIXES:
                ident = _sanitise_ident(alias or src.stem, taken_idents)
                _load_delimited_into_sqlite(conn, src, ident)
                sheet_map[ident] = {"path": str(src), "sheet": src.stem}
            else:
                openpyxl, derr = _require_openpyxl()
                if derr:
                    return derr
                try:
                    wb = openpyxl.load_workbook(filename=str(src), read_only=True, data_only=True)
                except Exception as e:
                    return _err("unsupported_format", f"Failed to open {path_str!r}: {e}")

                if sheet_filter:
                    if sheet_filter not in wb.sheetnames:
                        return _err(
                            "sheet_not_found",
                            f"Sheet {sheet_filter!r} not in {path_str!r}. Available: {wb.sheetnames}.",
                        )
                    target_sheets = [sheet_filter]
                else:
                    target_sheets = list(wb.sheetnames)

                for sheet_name in target_sheets:
                    ident_base = alias if (alias and len(target_sheets) == 1) else (
                        f"{alias}_{sheet_name}" if alias else sheet_name
                    )
                    ident = _sanitise_ident(ident_base, taken_idents)
                    _load_sheet_into_sqlite(conn, wb[sheet_name], ident)
                    sheet_map[ident] = {"path": str(src), "sheet": sheet_name}

        if not sheet_map:
            return _err("invalid_input", "No tables were loaded from the provided inputs.")

        try:
            cur = conn.execute(query)
        except sqlite3.Error as e:
            return _err("sql_error", f"{type(e).__name__}: {e}")

        cols = [d[0] for d in (cur.description or [])]
        max_rows = max(1, int(params.max_rows))
        rows: List[List[Any]] = []
        truncated = False
        for i, row in enumerate(cur):
            if i >= max_rows:
                truncated = True
                break
            rows.append([_jsonable(v) for v in row])

        return ToolResult(ok=True, data={
            "columns": cols,
            "rows": rows,
            "row_count": len(rows),
            "truncated": truncated,
            "tables": sheet_map,
        })
    finally:
        conn.close()


def _load_delimited_into_sqlite(conn: sqlite3.Connection, path: Path, ident: str) -> None:
    delim = _delimiter_for(path)
    with path.open("r", encoding="utf-8", newline="") as f:
        reader = csv.reader(f, delimiter=delim)
        rows = [row for row in reader]
    headers = rows[0] if rows else []
    data = rows[1:] if rows else []
    _create_and_fill_table(conn, ident, headers, data)


def _load_sheet_into_sqlite(conn: sqlite3.Connection, ws, ident: str) -> None:
    rows_iter = ws.iter_rows(values_only=True)
    try:
        first = next(rows_iter)
    except StopIteration:
        _create_and_fill_table(conn, ident, [], [])
        return
    headers = [str(_jsonable(v)) if v is not None else f"col_{i+1}" for i, v in enumerate(first)]
    data = ([_jsonable(v) for v in row] for row in rows_iter)
    _create_and_fill_table(conn, ident, headers, data)


def _create_and_fill_table(conn: sqlite3.Connection, ident: str, headers: List[Any], data) -> None:
    used: set = set()
    col_idents: List[str] = []
    for i, h in enumerate(headers):
        name = str(h) if h is not None and str(h).strip() else f"col_{i+1}"
        cleaned = _IDENT_RE.sub("_", name).strip("_") or f"col_{i+1}"
        if cleaned[0].isdigit():
            cleaned = "c_" + cleaned
        base = cleaned
        j = 2
        while cleaned.lower() in used:
            cleaned = f"{base}_{j}"
            j += 1
        used.add(cleaned.lower())
        col_idents.append(cleaned)

    if not col_idents:
        conn.execute(f'CREATE TABLE "{ident}" (col_1 TEXT)')
        return

    cols_sql = ", ".join(f'"{c}"' for c in col_idents)
    conn.execute(f'CREATE TABLE "{ident}" ({cols_sql})')
    placeholders = ",".join(["?"] * len(col_idents))
    insert_sql = f'INSERT INTO "{ident}" VALUES ({placeholders})'
    n = len(col_idents)

    def _normalised(rows_in):
        for r in rows_in:
            r = list(r)
            if len(r) < n:
                r = r + [None] * (n - len(r))
            elif len(r) > n:
                r = r[:n]
            yield r
    conn.executemany(insert_sql, _normalised(data))


# ---------------------------------------------------------------------------
# xlsx.write — create or overwrite a workbook from rows
# ---------------------------------------------------------------------------


def write(params, ctx: ToolCtx) -> ToolResult:
    out_suffix = Path(params.output).suffix.lower()
    if out_suffix in _XLSX_SUFFIXES:
        out_path, err = _check_output_path(params.output, params.overwrite, suffixes=_XLSX_SUFFIXES)
        if err:
            return err
        return _write_xlsx(out_path, params)
    if out_suffix in _DELIMITED_SUFFIXES:
        out_path, err = _check_output_path(params.output, params.overwrite, suffixes=_DELIMITED_SUFFIXES)
        if err:
            return err
        return _write_delimited(out_path, params)
    return _err(
        "unsupported_format",
        f"Output extension {out_suffix!r} not supported. Use .xlsx, .xlsm, .csv, or .tsv.",
    )


def _normalise_sheets_input(params) -> Tuple[Optional[Dict[str, dict]], Optional[ToolResult]]:
    """Return an ordered dict of {sheet_name: {headers, rows}}."""
    sheets = params.sheets
    if sheets is None:
        if params.headers is None and params.rows is None:
            return None, _err("invalid_input", "Provide either `sheets` or (`headers` and/or `rows`).")
        return {params.sheet_name or "Sheet1": {
            "headers": params.headers,
            "rows": params.rows or [],
        }}, None

    if not isinstance(sheets, dict) or not sheets:
        return None, _err("invalid_input", "`sheets` must be a non-empty object keyed by sheet name.")

    out: Dict[str, dict] = {}
    for name, spec in sheets.items():
        if isinstance(spec, dict):
            headers = spec.get("headers")
            rows = spec.get("rows") or []
        elif hasattr(spec, "headers"):  # pydantic _SheetSpec
            headers = spec.headers
            rows = spec.rows or []
        else:
            return None, _err("invalid_input", f"Sheet {name!r} must be an object with headers/rows.")
        out[name] = {"headers": headers, "rows": rows}
    return out, None


def _write_xlsx(out: Path, params) -> ToolResult:
    openpyxl, derr = _require_openpyxl()
    if derr:
        return derr

    sheets, err = _normalise_sheets_input(params)
    if err:
        return err

    wb = openpyxl.Workbook()
    # Remove the default empty sheet — we'll add our own in order.
    default = wb.active
    wb.remove(default)

    for sheet_name, spec in sheets.items():
        ws = wb.create_sheet(title=str(sheet_name)[:31])
        headers = spec.get("headers")
        if headers is not None:
            ws.append(list(headers))
        for row in spec.get("rows", []):
            ws.append(list(row))

    try:
        wb.save(str(out))
    except Exception as e:
        return _err("write_failed", f"Could not write {out}: {e}")

    return ToolResult(ok=True, data={
        "output": str(out),
        "sheet_names": list(sheets.keys()),
        "sheet_count": len(sheets),
    })


def _write_delimited(out: Path, params) -> ToolResult:
    sheets, err = _normalise_sheets_input(params)
    if err:
        return err
    if len(sheets) > 1:
        return _err(
            "invalid_input",
            "CSV/TSV outputs hold one sheet; pass a single sheet or write to .xlsx.",
        )
    name, spec = next(iter(sheets.items()))
    delim = _delimiter_for(out)
    try:
        with out.open("w", encoding="utf-8", newline="") as f:
            writer = csv.writer(f, delimiter=delim)
            if spec.get("headers") is not None:
                writer.writerow(list(spec["headers"]))
            for row in spec.get("rows", []):
                writer.writerow(list(row))
    except Exception as e:
        return _err("write_failed", f"Could not write {out}: {e}")
    return ToolResult(ok=True, data={
        "output": str(out),
        "sheet_names": [name],
        "sheet_count": 1,
    })


# ---------------------------------------------------------------------------
# xlsx.edit_cells — set individual cells in an existing workbook
# ---------------------------------------------------------------------------


def edit_cells(params, ctx: ToolCtx) -> ToolResult:
    src, err = _require_input(params.path, suffixes=_XLSX_SUFFIXES)
    if err:
        return err
    openpyxl, derr = _require_openpyxl()
    if derr:
        return derr

    out_path, oerr = _check_output_path(params.output, params.overwrite, suffixes=_XLSX_SUFFIXES)
    if oerr:
        return oerr

    try:
        wb = openpyxl.load_workbook(filename=str(src))
    except Exception as e:
        return _err("unsupported_format", f"Failed to open as xlsx: {e}")

    sheet_name = params.sheet or wb.sheetnames[0]
    if sheet_name not in wb.sheetnames:
        return _err(
            "sheet_not_found",
            f"Sheet {sheet_name!r} not in workbook. Available: {wb.sheetnames}.",
        )
    ws = wb[sheet_name]

    if not params.cells:
        return _err("invalid_input", "`cells` is empty; provide at least one {cell, value} entry.")

    written = 0
    formulas_present = False
    for entry in params.cells:
        if isinstance(entry, dict):
            cell_ref = entry.get("cell")
            value = entry.get("value")
        else:
            cell_ref = getattr(entry, "cell", None)
            value = getattr(entry, "value", None)
        if not cell_ref:
            return _err("invalid_input", f"Entry {entry!r} missing required field 'cell' (A1 notation).")
        try:
            ws[cell_ref] = value
        except Exception as e:
            return _err("invalid_input", f"Could not set {cell_ref!r}: {e}")
        if isinstance(value, str) and value.startswith("="):
            formulas_present = True
        written += 1

    try:
        wb.save(str(out_path))
    except Exception as e:
        return _err("write_failed", f"Could not write {out_path}: {e}")

    warnings: List[str] = []
    if formulas_present:
        warnings.append(
            "Formulas were written. Call xlsx.recalc to materialise their values "
            "before reading back with xlsx.read."
        )
    return ToolResult(
        ok=True,
        data={"output": str(out_path), "sheet": sheet_name, "cells_written": written},
        warnings=warnings,
    )


# ---------------------------------------------------------------------------
# xlsx.format — apply styling (font, fill, number format, width) to a range
# ---------------------------------------------------------------------------


def format_(params, ctx: ToolCtx) -> ToolResult:
    src, err = _require_input(params.path, suffixes=_XLSX_SUFFIXES)
    if err:
        return err
    openpyxl, derr = _require_openpyxl()
    if derr:
        return derr
    from openpyxl.styles import Font, PatternFill, Alignment  # type: ignore

    out_path, oerr = _check_output_path(params.output, params.overwrite, suffixes=_XLSX_SUFFIXES)
    if oerr:
        return oerr

    try:
        wb = openpyxl.load_workbook(filename=str(src))
    except Exception as e:
        return _err("unsupported_format", f"Failed to open as xlsx: {e}")

    sheet_name = params.sheet or wb.sheetnames[0]
    if sheet_name not in wb.sheetnames:
        return _err(
            "sheet_not_found",
            f"Sheet {sheet_name!r} not in workbook. Available: {wb.sheetnames}.",
        )
    ws = wb[sheet_name]

    if not params.range:
        return _err("invalid_input", "`range` is required (A1 notation, e.g. 'A1:C10' or 'B2').")

    try:
        cells_iter = ws[params.range]
    except Exception as e:
        return _err("invalid_input", f"Invalid range {params.range!r}: {e}")

    font_kwargs: Dict[str, Any] = {}
    if params.font_name is not None:
        font_kwargs["name"] = params.font_name
    if params.font_size is not None:
        font_kwargs["size"] = params.font_size
    if params.bold is not None:
        font_kwargs["bold"] = params.bold
    if params.italic is not None:
        font_kwargs["italic"] = params.italic
    if params.font_color is not None:
        font_kwargs["color"] = params.font_color

    fill = None
    if params.fill_color is not None:
        fill = PatternFill("solid", start_color=params.fill_color, end_color=params.fill_color)

    align = None
    if params.align is not None:
        align = Alignment(horizontal=params.align)

    def _apply(cell):
        if font_kwargs:
            cell.font = Font(**font_kwargs)
        if fill is not None:
            cell.fill = fill
        if align is not None:
            cell.alignment = align
        if params.number_format is not None:
            cell.number_format = params.number_format

    count = 0
    # ws[range] yields either a single cell, a row of cells, or a tuple of rows.
    if hasattr(cells_iter, "value"):  # single cell
        _apply(cells_iter)
        count = 1
    else:
        for row in cells_iter:
            if hasattr(row, "value"):
                _apply(row)
                count += 1
            else:
                for cell in row:
                    _apply(cell)
                    count += 1

    if params.column_widths:
        for col_letter, width in params.column_widths.items():
            ws.column_dimensions[col_letter].width = float(width)

    try:
        wb.save(str(out_path))
    except Exception as e:
        return _err("write_failed", f"Could not write {out_path}: {e}")

    return ToolResult(ok=True, data={
        "output": str(out_path),
        "sheet": sheet_name,
        "cells_formatted": count,
    })


# ---------------------------------------------------------------------------
# xlsx.convert — between xlsx/csv/tsv, with multi-sheet handling
# ---------------------------------------------------------------------------


def convert(params, ctx: ToolCtx) -> ToolResult:
    src, err = _require_input(params.path)
    if err:
        return err
    out_suffix = Path(params.output).suffix.lower()
    # In explode mode `output` is a directory, not a file — no extension required.
    if not params.explode_sheets and out_suffix not in _TABULAR_SUFFIXES:
        return _err("invalid_input", f"Output extension {out_suffix!r} not supported.")

    src_suffix = src.suffix.lower()

    # xlsx -> xlsx (sheet extract) or xlsx -> csv/tsv
    if src_suffix in _XLSX_SUFFIXES:
        openpyxl, derr = _require_openpyxl()
        if derr:
            return derr
        try:
            wb = openpyxl.load_workbook(filename=str(src), read_only=True, data_only=True)
        except Exception as e:
            return _err("unsupported_format", f"Failed to open as xlsx: {e}")

        if params.explode_sheets:
            return _explode_xlsx_to_csvs(wb, params, src)

        if out_suffix in _DELIMITED_SUFFIXES:
            sheet_name = params.sheet or wb.sheetnames[0]
            if sheet_name not in wb.sheetnames:
                return _err(
                    "sheet_not_found",
                    f"Sheet {sheet_name!r} not in workbook. Available: {wb.sheetnames}.",
                )
            out_path, oerr = _check_output_path(params.output, params.overwrite, suffixes=_DELIMITED_SUFFIXES)
            if oerr:
                return oerr
            ws = wb[sheet_name]
            delim = _delimiter_for(out_path)
            with out_path.open("w", encoding="utf-8", newline="") as f:
                writer = csv.writer(f, delimiter=delim)
                for row in ws.iter_rows(values_only=True):
                    writer.writerow([_jsonable(v) for v in row])
            return ToolResult(ok=True, data={
                "output": str(out_path),
                "sheet": sheet_name,
                "sheet_count": 1,
            })

        # xlsx -> xlsx (extract a sheet into its own workbook)
        out_path, oerr = _check_output_path(params.output, params.overwrite, suffixes=_XLSX_SUFFIXES)
        if oerr:
            return oerr
        sheet_name = params.sheet or wb.sheetnames[0]
        if sheet_name not in wb.sheetnames:
            return _err(
                "sheet_not_found",
                f"Sheet {sheet_name!r} not in workbook. Available: {wb.sheetnames}.",
            )
        new_wb = openpyxl.Workbook()
        new_wb.remove(new_wb.active)
        new_ws = new_wb.create_sheet(title=sheet_name[:31])
        for row in wb[sheet_name].iter_rows(values_only=True):
            new_ws.append([_jsonable(v) for v in row])
        new_wb.save(str(out_path))
        return ToolResult(ok=True, data={
            "output": str(out_path),
            "sheet": sheet_name,
            "sheet_count": 1,
        })

    # csv/tsv -> csv/tsv (delim change) or csv/tsv -> xlsx
    if out_suffix in _XLSX_SUFFIXES:
        openpyxl, derr = _require_openpyxl()
        if derr:
            return derr
        out_path, oerr = _check_output_path(params.output, params.overwrite, suffixes=_XLSX_SUFFIXES)
        if oerr:
            return oerr
        delim = _delimiter_for(src)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = src.stem[:31] or "Sheet1"
        with src.open("r", encoding="utf-8", newline="") as f:
            for row in csv.reader(f, delimiter=delim):
                ws.append(row)
        wb.save(str(out_path))
        return ToolResult(ok=True, data={
            "output": str(out_path),
            "sheet": ws.title,
            "sheet_count": 1,
        })

    # csv -> tsv or tsv -> csv
    out_path, oerr = _check_output_path(params.output, params.overwrite, suffixes=_DELIMITED_SUFFIXES)
    if oerr:
        return oerr
    src_delim = _delimiter_for(src)
    out_delim = _delimiter_for(out_path)
    with src.open("r", encoding="utf-8", newline="") as fin, \
            out_path.open("w", encoding="utf-8", newline="") as fout:
        writer = csv.writer(fout, delimiter=out_delim)
        for row in csv.reader(fin, delimiter=src_delim):
            writer.writerow(row)
    return ToolResult(ok=True, data={
        "output": str(out_path),
        "sheet": src.stem,
        "sheet_count": 1,
    })


def _explode_xlsx_to_csvs(wb, params, src: Path) -> ToolResult:
    out_dir = Path(params.output)
    if out_dir.suffix:
        out_dir = out_dir.parent / out_dir.stem
    out_dir.mkdir(parents=True, exist_ok=True)
    written: List[str] = []
    for name in wb.sheetnames:
        safe = re.sub(r"[^A-Za-z0-9._-]+", "_", name).strip("_") or "sheet"
        out_path = out_dir / f"{safe}.csv"
        if out_path.exists() and not params.overwrite:
            return _err("output_exists", f"{out_path} already exists. Pass overwrite=true.")
        with out_path.open("w", encoding="utf-8", newline="") as f:
            writer = csv.writer(f)
            for row in wb[name].iter_rows(values_only=True):
                writer.writerow([_jsonable(v) for v in row])
        written.append(str(out_path))
    return ToolResult(ok=True, data={
        "output_dir": str(out_dir),
        "files": written,
        "sheet_count": len(written),
    })


# ---------------------------------------------------------------------------
# xlsx.recalc — recalculate formulas via LibreOffice
# ---------------------------------------------------------------------------


_EXCEL_ERROR_TOKENS = ("#REF!", "#DIV/0!", "#VALUE!", "#N/A", "#NAME?", "#NUM!", "#NULL!")


def recalc(params, ctx: ToolCtx) -> ToolResult:
    src, err = _require_input(params.path, suffixes=_XLSX_SUFFIXES)
    if err:
        return err

    soffice = shutil.which("soffice") or shutil.which("libreoffice")
    if not soffice:
        return _err(
            "dependency_missing",
            "LibreOffice (`soffice`) is required to recalculate formulas but was not found on PATH.",
        )

    with tempfile.TemporaryDirectory() as tmpd:
        cmd = [
            soffice,
            "--headless",
            "--calc",
            "--convert-to", src.suffix.lstrip("."),
            "--outdir", tmpd,
            str(src),
        ]
        try:
            proc = subprocess.run(
                cmd,
                capture_output=True,
                timeout=max(5, int(params.timeout_seconds)),
            )
        except subprocess.TimeoutExpired:
            return _err("timeout", f"LibreOffice did not finish within {params.timeout_seconds}s.", retriable=True)
        except Exception as e:
            return _err("recalc_failed", f"Could not invoke LibreOffice: {e}")

        if proc.returncode != 0:
            return _err(
                "recalc_failed",
                f"LibreOffice exited {proc.returncode}: {proc.stderr.decode('utf-8', 'replace')[:400]}",
            )

        produced = Path(tmpd) / src.name
        if not produced.is_file():
            # LibreOffice sometimes renames; fall back to whatever single file landed.
            candidates = list(Path(tmpd).iterdir())
            if not candidates:
                return _err("recalc_failed", "LibreOffice produced no output file.")
            produced = candidates[0]

        dest_str = params.output or str(src)
        dest, derr = _check_output_path(
            dest_str,
            overwrite=(params.output is None) or params.overwrite,
            suffixes=_XLSX_SUFFIXES,
        )
        if derr:
            return derr
        shutil.copyfile(produced, dest)

    # Scan the recalculated workbook for residual Excel errors.
    openpyxl, derr2 = _require_openpyxl()
    if derr2:
        return derr2
    wb = openpyxl.load_workbook(filename=str(dest), data_only=True)
    error_summary: Dict[str, Dict[str, Any]] = {}
    total_errors = 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if isinstance(v, str) and v in _EXCEL_ERROR_TOKENS:
                    bucket = error_summary.setdefault(v, {"count": 0, "locations": []})
                    bucket["count"] += 1
                    if len(bucket["locations"]) < 20:
                        bucket["locations"].append(f"{sheet_name}!{cell.coordinate}")
                    total_errors += 1

    return ToolResult(ok=True, data={
        "output": str(dest),
        "status": "errors_found" if total_errors else "success",
        "total_errors": total_errors,
        "error_summary": error_summary,
    })
